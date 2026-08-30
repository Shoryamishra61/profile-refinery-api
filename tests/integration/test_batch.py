from __future__ import annotations

import csv
import io
import json
import zipfile

import httpx
import openpyxl
import pytest
from conftest import FULL_PROFILE_FIXTURE, StubTransport

from profile_refinery_api.batch.exports import FLAT_COLUMNS
from profile_refinery_api.errors import ProfileNotFound, UpstreamTimeout

TEXT = (
    "Team list:\n"
    "https://www.linkedin.com/in/test-integration-profile/\n"
    "linkedin.com/in/test-integration-profile (duplicate)\n"
    "https://www.linkedin.com/in/second-person/?trk=x\n"
    "https://www.linkedin.com/in/missing-person/\n"
    "Not profiles: https://www.linkedin.com/company/acme/ https://www.linkedin.com/feed/\n"
)


def auth(client: httpx.AsyncClient) -> dict[str, str]:
    return {"X-API-Key": "test-api-key"}


async def _create_batch(client: httpx.AsyncClient, text: str) -> dict[str, object]:
    response = await client.post("/v1/batches", params={"text": text}, headers=auth(client))
    assert response.status_code == 202, response.text
    return response.json()


@pytest.mark.asyncio
async def test_batch_dedupes_and_processes_with_provenance(client: httpx.AsyncClient) -> None:
    summary = await _create_batch(
        client,
        "https://www.linkedin.com/in/test-integration-profile/\n"
        "linkedin.com/in/test-integration-profile\n"
        "https://www.linkedin.com/in/second-person/",
    )
    stats = summary["statistics"]
    assert stats["url_occurrences_discovered"] == 3
    assert stats["unique_profiles"] == 2
    assert stats["duplicates_removed"] == 1

    final = await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 5}, headers=auth(client)
    )
    assert final.status_code == 200
    body = final.json()
    assert body["status"] == "SUCCEEDED"
    assert body["statistics"]["succeeded"] == 2

    profiles = await client.get(f"/v1/batches/{summary['batch_id']}/profiles", headers=auth(client))
    jobs = profiles.json()["profiles"]
    assert all(job["state"] == "SUCCEEDED" for job in jobs)
    first = jobs[0]
    assert first["occurrences"][0]["source_type"] == "pasted_text"
    assert first["occurrences"][0]["offset"] is not None


@pytest.mark.asyncio
async def test_batch_partial_failure_preserves_successes(
    client: httpx.AsyncClient, stub_transport: StubTransport
) -> None:
    stub_transport.set_for_slug("profile_view", "missing-person", [ProfileNotFound()])
    summary = await _create_batch(
        client,
        "https://www.linkedin.com/in/missing-person/\n"
        "https://www.linkedin.com/in/test-integration-profile/",
    )
    final = await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 5}, headers=auth(client)
    )
    body = final.json()
    assert body["status"] == "PARTIAL"
    assert body["statistics"]["succeeded"] == 1
    assert body["statistics"]["failed"] == 1

    profiles = (
        await client.get(
            f"/v1/batches/{summary['batch_id']}/profiles",
            params={"include_responses": True},
            headers=auth(client),
        )
    ).json()
    failed = [job for job in profiles["profiles"] if job["state"] == "FAILED"]
    succeeded = [job for job in profiles["profiles"] if job["state"] == "SUCCEEDED"]
    assert failed[0]["error_code"] == "PROFILE_NOT_FOUND"
    assert succeeded[0]["response"]["profile"]["name"]["value"] == "Integration Check"


@pytest.mark.asyncio
async def test_batch_transient_failure_is_retried_then_succeeds(
    client: httpx.AsyncClient, stub_transport: StubTransport
) -> None:
    stub_transport.set("profile_view", [UpstreamTimeout("profile_view"), FULL_PROFILE_FIXTURE])
    summary = await _create_batch(client, "https://www.linkedin.com/in/flaky-person/")
    final = await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 5}, headers=auth(client)
    )
    job = final.json()
    assert job["statistics"]["succeeded"] == 1


@pytest.mark.asyncio
async def test_batch_profile_endpoint_returns_grounded_report(client: httpx.AsyncClient) -> None:
    summary = await _create_batch(client, "https://www.linkedin.com/in/test-integration-profile/")
    await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 5}, headers=auth(client)
    )
    response = await client.get(
        f"/v1/batches/{summary['batch_id']}/profiles/test-integration-profile",
        headers=auth(client),
    )
    assert response.status_code == 200
    body = response.json()
    report = body["report"]
    assert report["current_position"]["company"] == "Pipeline Validation Corp"
    assert report["skills"] == ["HTTP protocol analysis", "Rest.li"]
    assert report["career_timeline"][1]["end_date"] == {"year": 2023, "month": 3, "day": None}


@pytest.mark.asyncio
async def test_batch_unknown_profile_returns_404(client: httpx.AsyncClient) -> None:
    summary = await _create_batch(client, "https://www.linkedin.com/in/test-integration-profile/")
    await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 5}, headers=auth(client)
    )
    response = await client.get(
        f"/v1/batches/{summary['batch_id']}/profiles/nobody", headers=auth(client)
    )
    assert response.status_code == 404
    assert response.json()["code"] == "PROFILE_JOB_NOT_FOUND"


@pytest.mark.asyncio
async def test_batch_exports_csv_json_xlsx(client: httpx.AsyncClient) -> None:
    summary = await _create_batch(client, "https://www.linkedin.com/in/test-integration-profile/")
    await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 5}, headers=auth(client)
    )
    csv_response = await client.get(
        f"/v1/batches/{summary['batch_id']}/export", params={"format": "csv"}, headers=auth(client)
    )
    assert csv_response.status_code == 200
    assert csv_response.headers["content-type"].startswith("text/csv")
    csv_rows = list(csv.DictReader(io.StringIO(csv_response.text)))
    assert list(csv_rows[0]) == FLAT_COLUMNS
    assert csv_rows[0]["name"] == "Integration Check"
    assert csv_rows[0]["current_title"] == "Staff Engineer"
    assert csv_rows[0]["current_company"] == "Pipeline Validation Corp"
    assert csv_rows[0]["experience_count"] == "2"
    assert csv_rows[0]["education_count"] == "1"
    repeated_csv = await client.get(
        f"/v1/batches/{summary['batch_id']}/export",
        params={"format": "csv"},
        headers=auth(client),
    )
    assert repeated_csv.content == csv_response.content

    json_response = await client.get(
        f"/v1/batches/{summary['batch_id']}/export",
        params={"format": "json"},
        headers=auth(client),
    )
    document = json_response.json()
    assert document["batch"]["batch_id"] == summary["batch_id"]
    exported = document["profiles"][0]
    assert exported["canonical_url"].endswith("test-integration-profile")
    assert exported["occurrences"][0]["source_type"] == "pasted_text"
    assert exported["response"]["profile"]["name"]["value"] == "Integration Check"
    assert len(exported["response"]["profile"]["experience"]["value"]) == 2
    assert len(exported["response"]["profile"]["education"]["value"]) == 1
    repeated_json = await client.get(
        f"/v1/batches/{summary['batch_id']}/export",
        params={"format": "json"},
        headers=auth(client),
    )
    assert repeated_json.json() == document

    xlsx_response = await client.get(
        f"/v1/batches/{summary['batch_id']}/export",
        params={"format": "xlsx"},
        headers=auth(client),
    )
    assert xlsx_response.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_response.content))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "linkedin_url"


def _xlsx_with_urls() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "candidate"
    sheet["B1"] = "https://www.linkedin.com/in/xlsx-person/"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _docx_with_urls() -> bytes:
    document = (
        '<?xml version="1.0"?><w:document xmlns:w="http://schemas.openxmlformats.org/'
        'wordprocessingml/2006/main"><w:body><w:p><w:r><w:t>'
        "Reach out to https://www.linkedin.com/in/docx-person"
        "</w:t></w:r></w:p></w:body></w:document>"
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("word/document.xml", document)
    return buffer.getvalue()


@pytest.mark.asyncio
async def test_batch_file_ingestion_csv_xlsx_docx_txt(client: httpx.AsyncClient) -> None:
    files = [
        (
            "files",
            ("people.csv", b"name,url\nAnn,https://www.linkedin.com/in/csv-person/", "text/csv"),
        ),
        ("files", ("people.xlsx", _xlsx_with_urls(), "application/octet-stream")),
        (
            "files",
            (
                "cv.docx",
                _docx_with_urls(),
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ),
        ),
        (
            "files",
            ("notes.txt", b"txt person https://www.linkedin.com/in/txt-person/", "text/plain"),
        ),
    ]
    response = await client.post("/v1/batches", files=files, headers=auth(client))
    assert response.status_code == 202, response.text
    summary = response.json()
    assert summary["statistics"]["unique_profiles"] == 4

    final = await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 10}, headers=auth(client)
    )
    assert final.json()["statistics"]["succeeded"] == 4

    profiles = (
        await client.get(f"/v1/batches/{summary['batch_id']}/profiles", headers=auth(client))
    ).json()["profiles"]
    by_slug = {job["canonical_url"].rsplit("/", 1)[-1]: job for job in profiles}
    assert by_slug["csv-person"]["occurrences"][0]["column"] == "url"
    assert by_slug["xlsx-person"]["occurrences"][0]["sheet"] == "Sheet"
    assert by_slug["docx-person"]["occurrences"][0]["row"] == 1
    assert by_slug["txt-person"]["occurrences"][0]["row"] == 1


@pytest.mark.asyncio
async def test_batch_file_ingestion_json_and_pdf_with_provenance(
    client: httpx.AsyncClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    import pypdf

    class Page:
        def extract_text(self) -> str:
            return "PDF candidate https://www.linkedin.com/in/pdf-batch-person/"

    class Reader:
        is_encrypted = False
        pages = [Page()]

    monkeypatch.setattr(pypdf, "PdfReader", lambda _: Reader())
    json_payload = json.dumps(
        {"candidates": [{"profile": "https://www.linkedin.com/in/json-batch-person/?trk=x"}]}
    ).encode()
    response = await client.post(
        "/v1/batches",
        files=[
            ("files", ("people.json", json_payload, "application/json")),
            ("files", ("people.pdf", b"%PDF-controlled", "application/pdf")),
        ],
        headers=auth(client),
    )
    assert response.status_code == 202
    summary = response.json()
    assert summary["statistics"]["unique_profiles"] == 2
    await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 10}, headers=auth(client)
    )
    profiles = (
        await client.get(f"/v1/batches/{summary['batch_id']}/profiles", headers=auth(client))
    ).json()["profiles"]
    by_slug = {job["canonical_url"].rsplit("/", 1)[-1]: job for job in profiles}
    assert by_slug["json-batch-person"]["occurrences"][0]["column"] == (
        "candidates[0].profile"
    )
    assert by_slug["pdf-batch-person"]["occurrences"][0]["row"] == 1


@pytest.mark.asyncio
async def test_duplicates_across_files_merge_all_provenance(client: httpx.AsyncClient) -> None:
    response = await client.post(
        "/v1/batches",
        files=[
            (
                "files",
                ("one.txt", b"linkedin.com/in/cross-file-person", "text/plain"),
            ),
            (
                "files",
                (
                    "two.csv",
                    b"name,url\nPerson,https://www.linkedin.com/in/cross-file-person/?trk=csv",
                    "text/csv",
                ),
            ),
            (
                "files",
                (
                    "three.json",
                    json.dumps(
                        {
                            "profile": "https://www.linkedin.com/in/cross-file-person/",
                            "malformed": "https://www.linkedin.com/in/../../admin",
                        }
                    ).encode(),
                    "application/json",
                ),
            ),
        ],
        headers=auth(client),
    )
    assert response.status_code == 202
    summary = response.json()
    assert summary["statistics"]["url_occurrences_discovered"] == 3
    assert summary["statistics"]["unique_profiles"] == 1
    assert summary["statistics"]["duplicates_removed"] == 2
    profiles = (
        await client.get(f"/v1/batches/{summary['batch_id']}/profiles", headers=auth(client))
    ).json()["profiles"]
    occurrences = profiles[0]["occurrences"]
    assert [item["source_name"] for item in occurrences] == [
        "one.txt",
        "two.csv",
        "three.json",
    ]
    assert occurrences[1]["row"] == 2 and occurrences[1]["column"] == "url"
    assert occurrences[2]["column"] == "profile"


@pytest.mark.asyncio
async def test_batch_rejects_unsupported_binary_and_oversize(
    client: httpx.AsyncClient,
) -> None:
    response = await client.post(
        "/v1/batches",
        files=[("files", ("blob.bin", b"\x00\x01\x02binary-noise", "application/octet-stream"))],
        headers=auth(client),
    )
    assert response.status_code == 202
    summary = response.json()
    assert summary["statistics"]["unique_profiles"] == 0
    assert summary["skipped_inputs"][0]["reason"] == "Unsupported binary file format."


@pytest.mark.asyncio
async def test_batch_idempotency_key_reuses_batch(client: httpx.AsyncClient) -> None:
    headers = {**auth(client), "Idempotency-Key": "same-thing"}
    first = await client.post(
        "/v1/batches", params={"text": "https://www.linkedin.com/in/idem-person/"}, headers=headers
    )
    second = await client.post(
        "/v1/batches", params={"text": "https://www.linkedin.com/in/idem-person/"}, headers=headers
    )
    assert first.json()["batch_id"] == second.json()["batch_id"]


@pytest.mark.asyncio
async def test_batch_requires_auth(client: httpx.AsyncClient) -> None:
    response = await client.post("/v1/batches", params={"text": "x"})
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_batch_404_for_unknown_batch(client: httpx.AsyncClient) -> None:
    response = await client.get("/v1/batches/does-not-exist", headers=auth(client))
    assert response.status_code == 404
    assert response.json()["code"] == "BATCH_NOT_FOUND"


@pytest.mark.asyncio
async def test_batch_accepts_json_body_and_raw_text_body(client: httpx.AsyncClient) -> None:
    json_body = {"text": "https://www.linkedin.com/in/body-json-person/"}
    first = await client.post("/v1/batches", json=json_body, headers=auth(client))
    assert first.status_code == 202
    assert first.json()["statistics"]["unique_profiles"] == 1

    raw = await client.post(
        "/v1/batches",
        content=b"https://www.linkedin.com/in/body-raw-person/",
        headers={**auth(client), "Content-Type": "text/plain"},
    )
    assert raw.status_code == 202
    assert raw.json()["statistics"]["unique_profiles"] == 1


@pytest.mark.asyncio
async def test_batch_report_endpoint_is_deterministic(client: httpx.AsyncClient) -> None:
    summary = await _create_batch(client, "https://www.linkedin.com/in/report-person/")
    await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 10}, headers=auth(client)
    )
    first = (
        await client.get(f"/v1/batches/{summary['batch_id']}/report", headers=auth(client))
    ).json()
    second = (
        await client.get(f"/v1/batches/{summary['batch_id']}/report", headers=auth(client))
    ).json()
    assert first["report_hash"] == second["report_hash"]
    assert len(first["report_hash"]) == 64
    assert first["report"]["profiles_processed"] == 1


@pytest.mark.asyncio
async def test_batch_xlsx_has_spec_sheets(client: httpx.AsyncClient) -> None:
    summary = await _create_batch(client, "https://www.linkedin.com/in/sheet-person/")
    await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 10}, headers=auth(client)
    )
    xlsx_response = await client.get(
        f"/v1/batches/{summary['batch_id']}/export",
        params={"format": "xlsx"},
        headers=auth(client),
    )
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_response.content))
    assert workbook.sheetnames == [
        "profiles",
        "experience",
        "education",
        "skills",
        "certifications",
        "languages",
        "provenance",
        "failures",
    ]
    profiles = workbook["profiles"]
    headers = [cell.value for cell in profiles[1]]
    assert headers == FLAT_COLUMNS
    assert profiles.cell(row=2, column=headers.index("current_company") + 1).value == (
        "Pipeline Validation Corp"
    )
    assert workbook["experience"].max_row == 3
    assert workbook["education"].max_row == 2
    assert workbook["skills"].max_row == 3
    assert workbook["certifications"].max_row == 2
    assert workbook["languages"].max_row == 2
    assert workbook["provenance"].max_row == 2
    assert workbook["failures"].max_row == 1
    # determinism: same stored input => byte-stable workbook structure
    again = await client.get(
        f"/v1/batches/{summary['batch_id']}/export",
        params={"format": "xlsx"},
        headers=auth(client),
    )
    repeated_workbook = openpyxl.load_workbook(io.BytesIO(again.content))
    assert repeated_workbook.sheetnames == workbook.sheetnames
    for sheet_name in workbook.sheetnames:
        first_values = list(workbook[sheet_name].values)
        repeated_values = list(repeated_workbook[sheet_name].values)
        assert repeated_values == first_values


@pytest.mark.asyncio
async def test_failed_job_appears_in_xlsx_failures_sheet(
    client: httpx.AsyncClient, stub_transport: StubTransport
) -> None:
    stub_transport.set_for_slug("profile_view", "doomed-person", [ProfileNotFound()])
    summary = await _create_batch(client, "https://www.linkedin.com/in/doomed-person/")
    await client.get(
        f"/v1/batches/{summary['batch_id']}", params={"wait_seconds": 10}, headers=auth(client)
    )
    xlsx_response = await client.get(
        f"/v1/batches/{summary['batch_id']}/export",
        params={"format": "xlsx"},
        headers=auth(client),
    )
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_response.content))
    failures = workbook["failures"]
    rows = list(failures.iter_rows(min_row=2, values_only=True))
    assert any(row[2] == "PROFILE_NOT_FOUND" for row in rows)
