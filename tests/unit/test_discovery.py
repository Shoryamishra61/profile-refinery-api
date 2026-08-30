from __future__ import annotations

import io
import json
import zipfile

import openpyxl
import pytest

from profile_refinery_api.batch.discovery import (
    canonicalize_post_url,
    dedupe,
    dedupe_posts,
    discover_in_text,
    discover_posts_in_text,
)
from profile_refinery_api.batch.ingest import (
    IngestError,
    ingest,
    ingest_links,
    sanitize_filename,
    sniff_kind,
)
from profile_refinery_api.errors import InvalidProfileUrl


def minimal_pdf_with_text(text: str) -> bytes:
    """Build a deterministic one-page PDF with extractable Helvetica text."""
    escaped = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    stream = f"BT /F1 10 Tf 10 100 Td ({escaped}) Tj ET".encode()
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        (
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 600 200] "
            b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>"
        ),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    document = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for number, body in enumerate(objects, start=1):
        offsets.append(len(document))
        document.extend(f"{number} 0 obj\n".encode() + body + b"\nendobj\n")
    xref = len(document)
    document.extend(f"xref\n0 {len(objects) + 1}\n".encode())
    document.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        document.extend(f"{offset:010d} 00000 n \n".encode())
    document.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode()
    )
    return bytes(document)


def test_discovers_bare_and_qualified_urls_with_provenance() -> None:
    text = "John: linkedin.com/in/john-smith\nnext https://www.linkedin.com/in/jane-doe/?trk=x end"
    found = discover_in_text(text, source_type="pasted_text")
    urls = [url for url, _ in found]
    assert urls == [
        "https://www.linkedin.com/in/john-smith",
        "https://www.linkedin.com/in/jane-doe",
    ]
    assert found[0][1].source_type == "pasted_text"
    assert found[0][1].offset == 6
    assert found[1][1].original_text == "https://www.linkedin.com/in/jane-doe/?trk=x"


def test_ignores_non_profile_linkedin_urls() -> None:
    text = (
        "company https://www.linkedin.com/company/acme/ "
        "jobs https://www.linkedin.com/jobs/view/123/ "
        "feed https://www.linkedin.com/feed/ person linkedin.com/in/real-person"
    )
    found = discover_in_text(text, source_type="pasted_text")
    assert [url for url, _ in found] == ["https://www.linkedin.com/in/real-person"]


def test_discovers_supported_post_urls_without_guessing_author() -> None:
    text = (
        "https://www.linkedin.com/posts/example-person_launch_activity-7471183910043922432-x "
        "https://www.linkedin.com/feed/update/urn:li:activity:7471183910043922432/?trk=x"
    )
    found = discover_posts_in_text(text, "pasted_text")
    assert [item[0] for item in found] == [
        "https://www.linkedin.com/posts/example-person_launch_activity-7471183910043922432-x",
        "https://www.linkedin.com/feed/update/urn:li:activity:7471183910043922432",
    ]
    assert found[0][1] is None
    assert found[1][1] == "urn:li:activity:7471183910043922432"
    assert found[0][2].offset == 0


def test_post_urls_are_deduped_with_occurrence_provenance() -> None:
    text = (
        "https://www.linkedin.com/feed/update/urn:li:activity:7471183910043922432\n"
        "https://linkedin.com/feed/update/urn:li:activity:7471183910043922432?trk=copy"
    )
    posts = dedupe_posts(discover_posts_in_text(text, "pasted_text"))
    assert len(posts) == 1
    assert len(posts[0].occurrences) == 2
    assert posts[0].activity_urn == "urn:li:activity:7471183910043922432"


def test_post_canonicalizer_rejects_non_post_surfaces() -> None:
    with pytest.raises(InvalidProfileUrl):
        canonicalize_post_url("https://www.linkedin.com/company/acme")


def test_ingest_links_finds_profiles_and_posts_in_one_file() -> None:
    payload = (
        b"profile,post\n"
        b"https://www.linkedin.com/in/file-person/,"
        b"https://www.linkedin.com/feed/update/urn:li:activity:7471183910043922432"
    )
    links = ingest_links(payload, "links.csv", "links.csv")
    assert links.profiles[0][0] == "https://www.linkedin.com/in/file-person"
    assert links.posts[0][0].endswith("urn:li:activity:7471183910043922432")
    assert links.posts[0][2].row == 2


def test_invalid_slug_never_becomes_a_profile() -> None:
    found = discover_in_text(
        "see linkedin.com/in/../../admin and linkedin.com/in/ok-person", "pasted_text"
    )
    assert [url for url, _ in found] == ["https://www.linkedin.com/in/ok-person"]


def test_dedupe_merges_occurrences_and_keeps_first_provenance() -> None:
    text = "a linkedin.com/in/dup b www.linkedin.com/in/dup c https://www.linkedin.com/in/dup/"
    occurrences = discover_in_text(text, source_type="pasted_text")
    profiles = dedupe(occurrences)
    assert len(profiles) == 1
    assert len(profiles[0].occurrences) == 3
    assert profiles[0].canonical.canonical_url == "https://www.linkedin.com/in/dup"


def test_sniff_kind_uses_content_not_extension() -> None:
    assert sniff_kind(b"%PDF-1.7 fake", "not-a-pdf.txt") == "pdf"
    workbook = openpyxl.Workbook()
    buffer = io.BytesIO()
    workbook.save(buffer)
    assert sniff_kind(buffer.getvalue(), "not-excel.txt") == "xlsx"
    assert sniff_kind(b"plain text https://www.linkedin.com/in/x", "a.txt") == "txt"
    assert sniff_kind(b'{"a": ["https://www.linkedin.com/in/x"]}', "b.txt") == "json"


def test_ooxml_zip_bomb_metadata_is_rejected(monkeypatch: pytest.MonkeyPatch) -> None:
    class Member:
        filename = "word/document.xml"
        file_size = 101 * 1024 * 1024

    class Archive:
        def __enter__(self) -> Archive:
            return self

        def __exit__(self, *args: object) -> None:
            return None

        def infolist(self) -> list[Member]:
            return [Member()]

    monkeypatch.setattr(zipfile, "ZipFile", lambda _: Archive())
    with pytest.raises(IngestError, match="expands beyond the safe limit"):
        sniff_kind(b"PK\x03\x04metadata-only", "resume.docx")


def test_txt_ingestion_records_line_numbers() -> None:
    payload = b"intro\ncontact linkedin.com/in/line-two\nlinkedin.com/in/line-three"
    found = ingest(payload, "people.txt", "people.txt")
    rows = {occ.row for _, occ in found}
    assert rows == {2, 3}


def test_csv_ingestion_records_row_and_column() -> None:
    payload = b"name,linkedin\nAnn,https://www.linkedin.com/in/ann-profile/\nBob,linkedin.com/in/bob-profile"
    found = ingest(payload, "people.csv", "people.csv")
    assert len(found) == 2
    assert found[0][1].column == "linkedin"
    assert found[0][1].row == 2


def test_xlsx_ingestion_records_sheet_and_cell() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Candidates"
    sheet["B2"] = "https://www.linkedin.com/in/sheet-person/"
    buffer = io.BytesIO()
    workbook.save(buffer)
    found = ingest(buffer.getvalue(), "people.xlsx", "people.xlsx")
    assert len(found) == 1
    occurrence = found[0][1]
    assert occurrence.sheet == "Candidates"
    assert occurrence.row == 2
    assert occurrence.column == "B2"


def test_docx_ingestion_records_paragraph() -> None:
    document = (
        '<?xml version="1.0"?><w:document xmlns:w="http://schemas.openxmlformats.org/'
        'wordprocessingml/2006/main"><w:body><w:p><w:r><w:t>first</w:t></w:r></w:p>'
        "<w:p><w:r><w:t>see https://www.linkedin.com/in/docx-person</w:t></w:r></w:p>"
        "</w:body></w:document>"
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("word/document.xml", document)
    found = ingest(buffer.getvalue(), "cv.docx", "cv.docx")
    assert len(found) == 1
    assert found[0][1].row == 2


def test_pdf_ingestion_records_page() -> None:
    payload = minimal_pdf_with_text("Profile https://www.linkedin.com/in/pdf-person/")
    found = ingest(payload, "profiles.pdf", "profiles.pdf")
    assert len(found) == 1
    assert found[0][0] == "https://www.linkedin.com/in/pdf-person"
    assert found[0][1].row == 1


def test_encrypted_pdf_is_rejected() -> None:
    pytest.importorskip("pypdf")
    from pypdf import PdfWriter

    writer = PdfWriter()
    writer.add_blank_page(width=200, height=200)
    writer.encrypt("owner-controlled-password")
    buffer = io.BytesIO()
    writer.write(buffer)
    with pytest.raises(IngestError, match="Encrypted PDFs are not supported"):
        ingest(buffer.getvalue(), "encrypted.pdf", "encrypted.pdf")


def test_pdf_page_limit_is_2000(monkeypatch: pytest.MonkeyPatch) -> None:
    import pypdf

    class Page:
        def __init__(self, text: str) -> None:
            self._text = text

        def extract_text(self) -> str:
            return self._text

    class Reader:
        is_encrypted = False
        pages = [Page("") for _ in range(1999)] + [
            Page("linkedin.com/in/page-2000"),
            Page("linkedin.com/in/page-2001"),
        ]

    monkeypatch.setattr(pypdf, "PdfReader", lambda _: Reader())
    found = ingest(b"%PDF-controlled", "limited.pdf", "limited.pdf")
    assert [url for url, _ in found] == ["https://www.linkedin.com/in/page-2000"]
    assert found[0][1].row == 2000


def test_json_ingestion_records_key_path() -> None:
    payload = json.dumps({"candidates": [{"url": "https://www.linkedin.com/in/json-person/"}]})
    found = ingest(payload.encode(), "data.json", "data.json")
    assert len(found) == 1
    assert found[0][1].column == "candidates[0].url"


def test_malformed_and_unsupported_files_are_explicit_errors() -> None:
    try:
        ingest(b"\x00\x01\x02binary", "blob.bin", "blob.bin")
    except IngestError:
        pass
    else:
        raise AssertionError("binary blob must raise IngestError")


@pytest.mark.parametrize(
    ("payload", "filename", "message"),
    [
        (b"PK\x03\x04broken", "broken.docx", "corrupt OOXML"),
        (b"%PDF-not-a-document", "broken.pdf", "PDF could not be parsed"),
    ],
)
def test_malformed_structured_files_are_explicit(
    payload: bytes, filename: str, message: str
) -> None:
    with pytest.raises(IngestError, match=message):
        ingest(payload, filename, filename)


def test_xlsx_sheet_and_row_limits_are_deterministic() -> None:
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "Included"
    first["A10000"] = "linkedin.com/in/row-10000"
    first["A10001"] = "linkedin.com/in/row-10001"
    for index in range(2, 22):
        workbook.create_sheet(f"Sheet{index}")
    workbook.worksheets[20]["A1"] = "linkedin.com/in/sheet-21"
    buffer = io.BytesIO()
    workbook.save(buffer)

    found = ingest(buffer.getvalue(), "limits.xlsx", "limits.xlsx")
    assert [url for url, _ in found] == ["https://www.linkedin.com/in/row-10000"]
    assert found[0][1].sheet == "Included"
    assert found[0][1].row == 10_000


def test_filenames_are_sanitized() -> None:
    assert sanitize_filename("../../etc/passwd") == "passwd"
    assert sanitize_filename("weird name?.xlsx") == "weird_name_.xlsx"
    assert sanitize_filename(None) == "upload"
